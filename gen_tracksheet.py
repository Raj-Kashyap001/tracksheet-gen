#!/usr/bin/env python3
"""Generate tracksheet Excel files from job order CSV reports."""

import hashlib
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Color palettes (header_fill, header_font, data_fills, border_color) ──────
PALETTES = [
    {"header_bg": "1F4E79", "header_fg": "FFFFFF", "row_alt": "D6E4F0", "border": "1F4E79", "title_bg": "2E75B6"},
    {"header_bg": "843C0C", "header_fg": "FFFFFF", "row_alt": "F2DCDB", "border": "843C0C", "title_bg": "C55A11"},
    {"header_bg": "1D6B3F", "header_fg": "FFFFFF", "row_alt": "D9E8D6", "border": "1D6B3F", "title_bg": "2E8B57"},
    {"header_bg": "4A1A6B", "header_fg": "FFFFFF", "row_alt": "E8D5F5", "border": "4A1A6B", "title_bg": "7B2D8E"},
    {"header_bg": "7B2D2D", "header_fg": "FFFFFF", "row_alt": "F2D5D5", "border": "7B2D2D", "title_bg": "A94442"},
    {"header_bg": "2D6B7B", "header_fg": "FFFFFF", "row_alt": "D5ECEF", "border": "2D6B7B", "title_bg": "3A9FBF"},
    {"header_bg": "6B5B1A", "header_fg": "FFFFFF", "row_alt": "F0ECD0", "border": "6B5B1A", "title_bg": "9B8B2A"},
    {"header_bg": "3B3B3B", "header_fg": "FFFFFF", "row_alt": "E0E0E0", "border": "3B3B3B", "title_bg": "5A5A5A"},
]

FONT_SANS = "Calibri"


def get_palette(account_name: str) -> dict:
    idx = int(hashlib.md5(account_name.encode()).hexdigest(), 16) % len(PALETTES)
    return PALETTES[idx]


def get_shift(now: datetime) -> str:
    hour = now.hour
    if 6 <= hour < 14:
        return "Shift A"
    elif 14 <= hour < 22:
        return "Shift B"
    else:
        return "Shift C"


def get_shift_date(now: datetime) -> str:
    shift = get_shift(now)
    fmt = "%-d %b %Y"
    if shift == "Shift C":
        prev = now - timedelta(days=1)
        return f"{prev.strftime(fmt)} and {now.strftime(fmt)}"
    return now.strftime(fmt)


def build_tracksheet(csv_path: Path, now: datetime) -> None:
    account_name = csv_path.stem
    shift = get_shift(now)
    shift_date = get_shift_date(now)
    palette = get_palette(account_name)

    # Read CSV
    df = pd.read_csv(csv_path, dtype=str)
    df.columns = df.columns.str.strip()

    # Normalize column names for flexible matching
    col_map = {c: c for c in df.columns}

    # Find required columns by partial match
    def find_col(options):
        for opt in options:
            for c in df.columns:
                if opt.lower() in c.lower():
                    return c
        return None

    col_route = find_col(["Trip/Route/Name", "Trip Route Name"])
    col_trip_id = find_col(["Trip id", "Trip ID", "trip_id"])
    col_vehicle = find_col(["Vehicle Number", "Vehicle No", "Vehicle_Number"])
    col_start_date = find_col(["Start Date", "StartDate"])
    col_start_time = find_col(["Start Time", "StartTime"])
    col_status = find_col(["TRIP STATUS", "Trip Status", "TRIP_STATUS"])
    col_driver_no = find_col(["Driver Number", "Driver_No"])

    if not all([col_route, col_trip_id, col_vehicle, col_status]):
        print(f"Warning: Missing required columns in {csv_path.name}, skipping.")
        print(f"  Found columns: {list(df.columns)}")
        return

    # Filter IN PROGRESS only
    df = df[df[col_status].str.upper().str.strip() == "IN PROGRESS"].copy()

    if df.empty:
        print(f"No IN PROGRESS trips found in {csv_path.name}, skipping.")
        return

    # Build Trip Start D&T
    if col_start_date and col_start_time:
        df["Trip Start D&T"] = df[col_start_date].fillna("").str.strip() + " " + df[col_start_time].fillna("").str.strip()
        df["Trip Start D&T"] = df["Trip Start D&T"].str.strip()
    elif col_start_date:
        df["Trip Start D&T"] = df[col_start_date].fillna("").str.strip()
    else:
        df["Trip Start D&T"] = ""

    # Sort by Trip Id
    df[col_trip_id] = pd.to_numeric(df[col_trip_id], errors="coerce")
    df = df.sort_values(by=col_trip_id, ascending=False).reset_index(drop=True)

    # ── Create workbook ──────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = account_name[:31]

    thin = Side(style="thin", color=palette["border"])
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)

    col_headers = [
        "S.No", "Loading From", "Trip Id", "Vehicle No.",
        "Trip Start D&T", "Trip Status", "Remarks",
        "",  # separator
        "Driver No.",
    ]
    num_cols = len(col_headers)

    # Column widths
    col_widths = [7, 35, 10, 16, 22, 14, 30, 3, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row 1: Title (merge only through Remarks column) ─────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws.cell(row=1, column=1, value=f"Tracksheet {account_name}")
    title_cell.font = Font(name=FONT_SANS, size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=palette["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    for c in range(2, 8):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", fgColor=palette["title_bg"])
        cell.border = border_all
    ws.cell(row=1, column=1).border = border_all

    # ── Row 2: Subheading (single row, two lines per cell) ──────────────
    sub_font = Font(name=FONT_SANS, size=12, bold=True)
    sub_fill = PatternFill("solid", fgColor=palette["row_alt"])
    sub_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sub_groups = [
        (1, 3, f"DATE:- {shift_date.upper()}\n{shift.upper()}"),
        (4, 5, "24 HOURS\nVEHICLES REPORT"),
        (6, 7, "REPORTING CYCLE\n6:00 AM TO 6:00 PM"),
    ]
    for sc, ec, text in sub_groups:
        ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
        cell = ws.cell(row=2, column=sc, value=text)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = sub_align

    ws.row_dimensions[2].height = 36

    # Borders on row 2
    for c in range(1, 8):  # only main table columns
        ws.cell(row=2, column=c).border = border_all

    # ── Row 3: Column headers ───────────────────────────────────────────
    header_font = Font(name=FONT_SANS, size=11, bold=True, color=palette["header_fg"])
    header_fill = PatternFill("solid", fgColor=palette["header_bg"])
    for c, hdr in enumerate(col_headers, 1):
        cell = ws.cell(row=3, column=c, value=hdr)
        if c <= 7:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border_all
        else:
            cell.font = Font(name=FONT_SANS, size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 24

    # ── Data rows ────────────────────────────────────────────────────────
    data_font = Font(name=FONT_SANS, size=10)
    data_fill_odd = PatternFill("solid", fgColor="FFFFFF")
    data_fill_even = PatternFill("solid", fgColor=palette["row_alt"])

    # Approximate chars-per-line per column (based on col_widths and font size 10)
    # Each col_width unit ~ 7px, font size 10 ~ 7px wide per char → ~1 char per width unit
    chars_per_line = {i: max(1, int(col_widths[i - 1] * 1.1)) for i in range(1, num_cols + 1)}
    line_height = 15  # pixels per text line

    for idx, (_, row) in enumerate(df.iterrows()):
        r = 4 + idx
        fill = data_fill_even if idx % 2 == 0 else data_fill_odd
        values = [
            idx + 1,
            str(row.get(col_route, "")).strip(),
            int(row[col_trip_id]) if pd.notna(row[col_trip_id]) else "",
            str(row.get(col_vehicle, "")).strip(),
            str(row.get("Trip Start D&T", "")).strip(),
            "",  # Trip Status — formula set below
            "",  # Remarks
            "",  # separator
            str(row.get(col_driver_no, "")).strip() if col_driver_no else "",
        ]
        max_lines = 1
        for c, val in enumerate(values, 1):
            text = str(val)
            lines_needed = max(1, -(-len(text) // chars_per_line[c]))
            max_lines = max(max_lines, lines_needed)
            cell = ws.cell(row=r, column=c, value=val)
            if c <= 7:  # main table columns
                cell.font = data_font
                cell.fill = fill
                cell.border = border_all
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="center",
                    horizontal="center",
                )
            else:  # extra columns — plain, no formatting
                cell.font = Font(name=FONT_SANS, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Trip Status formula: check Remarks (col7) for "trip end"
        ws.cell(row=r, column=6).value = f'=IF(ISNUMBER(SEARCH("trip end",G{r})),"COMPLETED","IN PROGRESS")'

        ws.row_dimensions[r].height = max(20, max_lines * line_height)

    # ── Save ─────────────────────────────────────────────────────────────
    out_dir = Path("out")
    out_dir.mkdir(exist_ok=True)
    date_str = now.strftime("%d-%m-%Y")
    out_file = out_dir / f"{account_name} {date_str} {shift} Tracksheet.xlsx"
    wb.save(out_file)
    print(f"Generated: {out_file}")


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: gen_tracksheet.py <file1.csv> [file2.csv ...]")
        sys.exit(1)

    now = datetime.now()
    for arg in sys.argv[1:]:
        csv_path = Path(arg)
        if not csv_path.exists():
            print(f"File not found: {csv_path}")
            continue
        build_tracksheet(csv_path, now)


if __name__ == "__main__":
    main()
